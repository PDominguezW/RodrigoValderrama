import openpyxl
import json
import time

def calculate_score(rut, data):

    # Read 'modelo_evaluacion.xlsx' with openpyxl
    workbook = openpyxl.load_workbook('modelo_evaluacion.xlsx')

    # Get the sheet you want to edit
    sheet = workbook['1.Datos']

    # Search for the B row that contains rut
    rut_formateado = rut
    rut_sin_verificador = rut_formateado.split('-')[0]

    # Read data from 'Tabla_SII.xlsx'
    sheet_sii = openpyxl.load_workbook('Tabla_SII.xlsx', read_only=True)['Tabla_SII']

    # Search for the B row that contains rut_sin_verificador
    encontrado = False
    for row in sheet_sii.iter_rows(min_row=2, max_col=23):
        if str(row[1].value) == str(rut_sin_verificador):
            valores_sii = row
            encontrado = True
            break

    # Enter rut
    sheet['B2'] = rut

    if encontrado:
        print(f"Score Calculator: Rut encontrado en Tabla_SII.xlsx")
        sheet['B6'] = rut.split('-')[1]

        inicio_actividades = str(valores_sii[7].value)
        if inicio_actividades.count('-') == 2:
            inicio_actividades_formateado = inicio_actividades.split(' ')[0].replace('-', '/')
            sheet['B7'] = inicio_actividades_formateado

            inicio_actividades_year = int(inicio_actividades_formateado.split('/')[0])
            actual_year = int(time.strftime("%Y"))
            sheet['B14'] = str(int(actual_year - inicio_actividades_year)) + " Años"
        else:
            sheet['B7'] = None
            sheet['B14'] = None

            inicio_actividades_formateado = inicio_actividades

        tamano_codigo = int(valores_sii[4].value)
        sheet['B8'] = tamano_codigo
        # Fecha hoy B9, no se modifica
        sheet['B10'] = str(rut)
        sheet['B11'] = str(valores_sii[3].value)
        sheet['B12'] = str(valores_sii[8].value)
        sheet['B13'] = str(valores_sii[5].value)

        trabajadores_valor = valores_sii[6].value
        if str(trabajadores_valor).isnumeric():
            trabajadores = int(trabajadores_valor)
            if trabajadores < 5:
                sheet['B15'] = "< 5 Trabajadores"
            else:
                sheet['B15'] = str(trabajadores_valor) + " Trabajadores"

        # Consideramos tamaño en B8
        sheet['B16'] = sheet[f'F{tamano_codigo + 6}'].value
        sheet['B17'] = valores_sii[17].value
        sheet['B18'] = sheet[f'J{tamano_codigo + 6}'].value
    else:
        print(f"Score Calculator: Rut no encontrado")

    # INGRESAMOS DATA EXPERIAN
    sheet['B21'] = data["experian"]["resumen_avaluo_bienes_raices"]["total_protestos_y_documentos"]
    sheet['B22'] = data["experian"]["resumen_avaluo_bienes_raices"]["total_en_pesos"]
    sheet['B23'] = data["experian"]["resumen_morosidad"]["nro_acreedores"]
    sheet['B24'] = data["experian"]["resumen_morosidad"]["total_pesos"]
    sheet['B25'] = data["experian"]["resumen_morosidad"]["total_doc_impagos"]

    if data["experian"]["resumen_socios_sociedades"]["rut_socio"]:
        sheet['B28'] = data["experian"]["resumen_socios_sociedades"]["rut_socio"].split('-')[0]
        sheet['B29'] = data["experian"]["resumen_socios_sociedades"]["rut_socio"].split('-')[1]
    else:
        sheet['B28'] = "NULL"
        sheet['B29'] = "NULL"

    sheet['B30'] = data["experian"]["resumen_socios_sociedades"]["data"]["resumen_avaluo_bienes_raices"]["total_protestos_y_documentos"]
    sheet['B31'] = data["experian"]["resumen_socios_sociedades"]["data"]["resumen_avaluo_bienes_raices"]["total_en_pesos"]
    sheet['B32'] = data["experian"]["resumen_socios_sociedades"]["data"]["resumen_morosidad"]["nro_acreedores"]
    sheet['B33'] = data["experian"]["resumen_socios_sociedades"]["data"]["resumen_morosidad"]["total_pesos"]
    sheet['B34'] = data["experian"]["resumen_socios_sociedades"]["data"]["resumen_morosidad"]["total_doc_impagos"]

    # INGRESAMOS DATA DEALERNET
    periodos = list(data["dealernet"]["empresa"]["AL DIA E IMPAGOS <30 DIAS"].keys())
    periodos_letras = ['B', 'C', 'D', 'E']

    sheet['B40'] = periodos[0]
    sheet['C40'] = periodos[1]
    sheet['D40'] = periodos[2]
    sheet['E40'] = periodos[3]

    sheet['B65'] = periodos[0]
    sheet['C65'] = periodos[1]
    sheet['D65'] = periodos[2]
    sheet['E65'] = periodos[3]

    columnas = [
        'AL DIA E IMPAGOS <30 DIAS',
        'IMPAGOS 30 Y 90 DIAS',
        'IMPAGOS 90 Y 180 DIAS',
        'IMPAGOS 180 DIAS Y 3 ANOS',
        'IMPAGOS >= 3 ANOS',
        'CREDITOS DE CONSUMO',
        'NRO. ENTIDADES CRED.CONSUMO',
        'CREDITOS PARA VIVIENDA',
        'OPERACIONES FINANCIERAS',
        'INSTRUM. DEUDAS ADQUIRIDOS',
        'CREDITOS COMERCIALES',
        'DEUDA COM. VIGENTE MEX',
        'DEUDA COM. VENCIDA MEX',
        'INDIRECTA IMPAGOS <30 DIAS',
        'INDIRECTA IMPAGOS 30 DIAS Y 3 ANOS',
        'INDIRECTA IMPAGOS >= 3 ANOS',
        'LINEA CREDITO DISPONIBLE',
        'CREDITOS CONTINGENTES',
        'NRO. ENTIDADES.CRED.COM ER.',
        'CREDITOS LEASING AL DIA',
        'CREDITOS LEASING IMPAGO'
    ]

    # Dealernet empresa
    for columna in columnas:
        for i in range(4):
            valor_data = data["dealernet"]["empresa"][columna][periodos[i]]
            valor_formateado = int(valor_data)
            sheet[f'{periodos_letras[i]}{columnas.index(columna) + 41}'] = valor_formateado

    if data["experian"]["resumen_socios_sociedades"]["rut_socio"]:
        # Dealernet socio
        for columna in columnas:
            for i in range(4):
                valor_data = data["dealernet"]["socio"][columna][periodos[i]]
                valor_formateado = int(valor_data)
                sheet[f'{periodos_letras[i]}{columnas.index(columna) + 66}'] = valor_formateado

    # Save the changes to the workbook
    workbook.save('modelo_evaluacion_result.xlsx')

    return 'modelo_evaluacion_result.xlsx'

if __name__ == "__main__":

    # Read the file 'new_data.json'
    with open('new_data.json') as json_file:
        data = json.load(json_file)

    # Run Flask app 76109342
    calculate_score("50026400-4", data)